"""Importación y exportación de preguntas de examen en Excel (openpyxl)."""
from io import BytesIO

from openpyxl import Workbook, load_workbook

from quizzes.models import ThemeExam, ExamQuestion, ExamAnswerOption

HEADER_ALIASES = {
    'pregunta': 'pregunta',
    'opcion_a': 'opcion_a',
    'opcion a': 'opcion_a',
    'opcion_b': 'opcion_b',
    'opcion b': 'opcion_b',
    'opcion_c': 'opcion_c',
    'opcion c': 'opcion_c',
    'correcta': 'correcta',
    'explicacion': 'explicacion',
    'explicación': 'explicacion',
}

EXPORT_HEADERS = (
    'pregunta',
    'opcion_a',
    'opcion_b',
    'opcion_c',
    'correcta',
    'explicacion',
)


def _normalize_header(cell_value):
    if cell_value is None:
        return ''
    s = str(cell_value).strip().lower().replace('ó', 'o')
    return HEADER_ALIASES.get(s, s.replace(' ', '_'))


def _parse_correct_cell(val):
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ('a', '1'):
        return 0
    if s in ('b', '2'):
        return 1
    if s in ('c', '3'):
        return 2
    return None


def parse_excel_rows(file_obj):
    """
    Lee un .xlsx y devuelve (rows, errors). Cada fila incluye opcionalmente 'explicacion'.
    """
    errors = []
    rows_out = []
    try:
        wb = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as e:
        return [], [f'No se pudo leer el archivo Excel: {e}']

    ws = wb.active
    iter_rows = ws.iter_rows(min_row=1, values_only=True)
    try:
        header_row = next(iter_rows)
    except StopIteration:
        return [], ['El archivo está vacío.']

    headers = []
    for c in header_row:
        headers.append(_normalize_header(c))

    try:
        idx_p = headers.index('pregunta')
        idx_a = headers.index('opcion_a')
        idx_b = headers.index('opcion_b')
        idx_c = headers.index('opcion_c')
        idx_ok = headers.index('correcta')
    except ValueError:
        return [], [
            'La primera fila debe incluir al menos: '
            'pregunta | opcion_a | opcion_b | opcion_c | correcta '
            '(y opcionalmente explicacion).'
        ]

    try:
        idx_exp = headers.index('explicacion')
    except ValueError:
        idx_exp = None

    max_idx = max(idx_p, idx_a, idx_b, idx_c, idx_ok, idx_exp or 0)

    for line_no, row in enumerate(iter_rows, start=2):
        if row is None or all(v is None or str(v).strip() == '' for v in row):
            continue
        cells = list(row) + [None] * max(0, max_idx + 1 - len(row))
        pregunta = cells[idx_p] if idx_p < len(cells) else None
        a = cells[idx_a] if idx_a < len(cells) else None
        b = cells[idx_b] if idx_b < len(cells) else None
        c = cells[idx_c] if idx_c < len(cells) else None
        cor = cells[idx_ok] if idx_ok < len(cells) else None
        exp_cell = cells[idx_exp] if idx_exp is not None and idx_exp < len(cells) else None

        pregunta = (str(pregunta).strip() if pregunta is not None else '')
        a = (str(a).strip() if a is not None else '')
        b = (str(b).strip() if b is not None else '')
        c_text = (str(c).strip() if c is not None else '')
        expl = (str(exp_cell).strip() if exp_cell is not None else '') if idx_exp is not None else ''

        if not pregunta and not a and not b and not c_text:
            continue

        ci = _parse_correct_cell(cor)
        if not pregunta:
            errors.append(f'Fila {line_no}: falta el texto de la pregunta.')
            continue
        if not a or not b or not c_text:
            errors.append(f'Fila {line_no}: las tres opciones son obligatorias.')
            continue
        if ci is None:
            errors.append(
                f'Fila {line_no}: «correcta» debe ser a, b, c (o 1, 2, 3).'
            )
            continue

        row_dict = {
            'pregunta': pregunta,
            'opcion_a': a,
            'opcion_b': b,
            'opcion_c': c_text,
            'correct_index': ci,
            'explicacion': expl,
        }
        rows_out.append(row_dict)

    wb.close()
    return rows_out, errors


def import_questions_for_exam(exam: ThemeExam, rows, replace_existing: bool):
    """Crea preguntas desde filas parseadas. Usar dentro de transaction.atomic."""
    if replace_existing:
        exam.questions.all().delete()

    order_base = 0
    if not replace_existing:
        last = ExamQuestion.objects.filter(exam=exam).order_by('-order').first()
        if last:
            order_base = last.order + 1

    for i, row in enumerate(rows):
        q = ExamQuestion.objects.create(
            exam=exam,
            text=row['pregunta'],
            order=order_base + i,
            correct_explanation=row.get('explicacion') or '',
        )
        texts = [row['opcion_a'], row['opcion_b'], row['opcion_c']]
        ci = row['correct_index']
        for j, t in enumerate(texts):
            ExamAnswerOption.objects.create(
                question=q,
                text=t,
                is_correct=(j == ci),
            )


def export_exam_to_workbook(exam: ThemeExam) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = 'preguntas'
    ws.append(list(EXPORT_HEADERS))
    questions = exam.questions.prefetch_related('answer_options').order_by('order', 'id')
    for q in questions:
        opts = list(q.answer_options.order_by('id'))
        if len(opts) != 3:
            continue
        correct_letter = None
        for idx, o in enumerate(opts):
            if o.is_correct:
                correct_letter = ('a', 'b', 'c')[idx]
                break
        if not correct_letter:
            continue
        ws.append([
            q.text,
            opts[0].text,
            opts[1].text,
            opts[2].text,
            correct_letter,
            (q.correct_explanation or '').strip(),
        ])
    return wb


def build_template_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = 'preguntas'
    ws.append(list(EXPORT_HEADERS))
    ws.append([
        'Ejemplo: ¿Cuál es la capital de Francia?',
        'Lyon',
        'París',
        'Marsella',
        'b',
        'París es la capital; Lyon y Marsella son grandes ciudades pero no la capital.',
    ])
    return wb


def workbook_to_response(wb: Workbook, filename: str):
    from django.http import HttpResponse

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
